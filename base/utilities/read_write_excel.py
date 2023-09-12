import openpyxl
import xlsxwriter


class ReadFromExcel(object):

    @staticmethod
    def readFromExcel(testdatapath):
        Dict = {}
        book = openpyxl.load_workbook(testdatapath)
        sheet = book.active
        for i in range(1, sheet.max_column+1):  # to get column
            values = []
            if sheet.cell(row=1, column=i).value not in Dict:
                Dict[sheet.cell(row=1, column=i).value] = list()
            for j in range(2, sheet.max_row+1):   # to get rows
                if sheet.cell(row=j, column=i).value is not None:
                    values.append(sheet.cell(row=j, column=i).value)
            Dict[sheet.cell(row=1, column=i).value].extend(values)
        return Dict


class WriteOnExcel(object):

    @staticmethod
    def write_on_excel(testdatapath,content):
        workbook = xlsxwriter.Workbook(testdatapath)
        worksheet = workbook.add_worksheet()
        column = 0
        for i in range(len(content)):
            row = 0
            # iterating through content list
            for item in content[i]:
                worksheet.write(row, column, item)
                row += 1
            column+=1
        workbook.close()
